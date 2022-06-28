from array import array
from multiprocessing.sharedctypes import RawArray
from time import sleep
from typing import final
# from openpyxl import load_workbook
import openpyxl
import os
import glob
import mymodule


lines = os.listdir('//ALYANS/ovk/Arhiv/')
linsize = len(lines)
summ = 0


for name_line in lines:
    all_papki_prover = os.listdir(f'//ALYANS/ovk/Arhiv/{name_line}/')
    summ += 1
    my_print = f'{summ} из {linsize} - {name_line}'
    print(my_print)

    for i in all_papki_prover:

        if 'Elements' in all_papki_prover:

            all_papki = os.listdir(f'//ALYANS/ovk/Arhiv/{name_line}/Elements/')

            for name_papka in all_papki:
                all_file_in_papka = os.listdir(f'//ALYANS/ovk/Arhiv/{name_line}/Elements/{name_papka}/') 

                for name_exel in all_file_in_papka:
                    if name_exel.endswith(".xlsx"):

                        arr_name_exels = []
                        arr_name_exels.append(name_exel)

                        for name_exel in arr_name_exels:

                            way_to_exel = f'//ALYANS/ovk/Arhiv/{name_line}/Elements/{name_papka}/{name_exel}'

                            wb = openpyxl.reader.excel.load_workbook(filename=way_to_exel)

                            wb.active
                            sheet = wb.active

                            for i in range(27, 57) :
                                a = sheet['C' + str(i)].value

                                if a == 'Наименование и номер документа, удостоверяющего качество: ':
                                    break
                                else:
                                    for i2 in range(27, 98):
                                        if sheet['G' + str(i2)].value == sheet['R' + str(i)].value:
                                        
                                            zaivka = mymodule.dell_slo(sheet['C23'].value)
                                            por_act = mymodule.dell_slo_o(name_papka)

                                            tab = [name_line, por_act, sheet['C2'].value, zaivka, sheet['C' + str(i)].value, sheet['F' + str(i)].value,sheet['R' + str(i)].value,sheet['U' + str(i)].value,sheet['X' + str(i)].value, sheet['K' + str(i2)].value, sheet['X' + str(i2)].value]
                                            

                                            fn = 'C:/Users/user/Desktop/Svod/Свод.xlsx'
                                            wb = openpyxl.load_workbook(fn)
                                            ws = wb['Лист1']
                                            ws.append(tab)

                                            wb.save(fn)
                                            wb.close
